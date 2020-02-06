#!/bin/python3

import openpyxl

# 새로 엑셀파일 만들기
class MakeExcel():
    def __init__(self, fileName):
        self.fileName = fileName + ".xlsx"
    
    def newfile(self):
        
        wb = openpyxl.Workbook() 
        Sheet_name = wb.sheetnames
        wb.save(filename=self.fileName)
        wb.close()

    def createsheet(self,sheetName):

        opwb = openpyxl.load_workbook(self.fileName)
        opwb.create_sheet(sheetName,index=-1)
        opwb.save(filename=self.fileName)
        opwb.close()

    def insertdata(self,sheetName,cell,val,mode=''):
        
        opwb = openpyxl.load_workbook(self.fileName)
        opws = opwb[sheetName]
        if mode == '':
            opws[cell] = val

        else:
            opws.append(val)
        opwb.save(filename=self.fileName)
        opwb.close()