#!/bin/python3

import openpyxl

# 새로 엑셀파일 만들기

#wb = openpyxl.Workbook() 
#Sheet_name = wb.sheetnames
#wb.save(filename='Test.xlsx')

# File Check
opwb = openpyxl.Workbook('Test.xlsx')
#sheet1 = opwb.active
sheet1 = opwb.create_sheet("Mysheet")
sheet1.title = 'abc'
opwb.save(filename='Test.xlsx')

print (opwb.sheetnames)